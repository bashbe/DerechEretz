import os
from datetime import datetime

from flask import current_app
from openpyxl import Workbook
from xhtml2pdf import pisa

from app.extensions import db
from app.models import Eleve, InfractionMineure, Note, Presence, RapportGenere
from app.services import calculer_moyenne_generale, calculer_moyenne_matiere
from app.models import Matiere


def _dossier_rapports():
    dossier = os.path.join(current_app.instance_path, "rapports")
    os.makedirs(dossier, exist_ok=True)
    return dossier


def _horodatage():
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")


def _ecrire_pdf(html, chemin):
    with open(chemin, "wb") as fichier:
        pisa.CreatePDF(html, dest=fichier)


def generer_rapport_notes(classe, trimestre):
    matieres = Matiere.query.order_by(Matiere.nom).all()
    lignes = []
    for eleve in sorted(classe.eleves, key=lambda e: e.nom):
        moyennes_matieres = {
            m.nom: calculer_moyenne_matiere(eleve.id, m.id, trimestre) for m in matieres
        }
        moyenne_generale = calculer_moyenne_generale(eleve.id, trimestre)
        lignes.append((eleve, moyennes_matieres, moyenne_generale))

    titre = f"Notes {classe.nom} - {trimestre}"
    horodatage = _horodatage()

    html = _render_html_notes(classe, trimestre, matieres, lignes)
    chemin_pdf = os.path.join(_dossier_rapports(), f"notes_{classe.id}_{trimestre}_{horodatage}.pdf")
    _ecrire_pdf(html, chemin_pdf)

    chemin_excel = os.path.join(
        _dossier_rapports(), f"notes_{classe.id}_{trimestre}_{horodatage}.xlsx"
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["Élève"] + [m.nom for m in matieres] + ["Moyenne générale"])
    for eleve, moyennes_matieres, moyenne_generale in lignes:
        ws.append(
            [eleve.nom_complet]
            + [moyennes_matieres[m.nom] if moyennes_matieres[m.nom] is not None else "" for m in matieres]
            + [moyenne_generale if moyenne_generale is not None else ""]
        )
    wb.save(chemin_excel)

    rapport = RapportGenere(
        type="notes",
        titre=titre,
        fichier_pdf=chemin_pdf,
        fichier_excel=chemin_excel,
    )
    db.session.add(rapport)
    db.session.commit()
    return rapport


def _render_html_notes(classe, trimestre, matieres, lignes):
    lignes_html = ""
    for eleve, moyennes_matieres, moyenne_generale in lignes:
        cellules = "".join(
            f"<td>{moyennes_matieres[m.nom] if moyennes_matieres[m.nom] is not None else '-'}</td>"
            for m in matieres
        )
        lignes_html += (
            f"<tr><td>{eleve.nom_complet}</td>{cellules}"
            f"<td><b>{moyenne_generale if moyenne_generale is not None else '-'}</b></td></tr>"
        )
    entetes = "".join(f"<th>{m.nom}</th>" for m in matieres)
    return f"""
    <html><body>
    <h2>Rapport de notes — {classe.nom} — {trimestre}</h2>
    <table border="1" cellpadding="4" cellspacing="0" width="100%">
      <tr><th>Élève</th>{entetes}<th>Moyenne générale</th></tr>
      {lignes_html}
    </table>
    </body></html>
    """


def generer_rapport_absences(classe, date_debut, date_fin):
    lignes = []
    for eleve in sorted(classe.eleves, key=lambda e: e.nom):
        entrees = Presence.query.filter(
            Presence.eleve_id == eleve.id,
            Presence.statut != "present",
            Presence.date >= date_debut,
            Presence.date <= date_fin,
        ).all()
        justifiees = sum(1 for p in entrees if p.statut == "absent" and p.justifie)
        injustifiees = sum(1 for p in entrees if p.statut == "absent" and not p.justifie)
        retards = sum(1 for p in entrees if p.statut == "retard")
        lignes.append((eleve, justifiees, injustifiees, retards))

    periode_str = f"{date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    titre = f"Absences {classe.nom} — {periode_str}"
    horodatage = _horodatage()

    lignes_html = "".join(
        f"<tr><td>{eleve.nom_complet}</td><td>{j}</td><td>{ij}</td><td>{r}</td></tr>"
        for eleve, j, ij, r in lignes
    )
    html = f"""
    <html><body>
    <h2>Rapport d'absences — {classe.nom} — {periode_str}</h2>
    <table border="1" cellpadding="4" cellspacing="0" width="100%">
      <tr><th>Élève</th><th>Absences justifiées</th><th>Absences injustifiées</th><th>Retards</th></tr>
      {lignes_html}
    </table>
    </body></html>
    """
    chemin_pdf = os.path.join(_dossier_rapports(), f"absences_{classe.id}_{horodatage}.pdf")
    _ecrire_pdf(html, chemin_pdf)

    chemin_excel = os.path.join(_dossier_rapports(), f"absences_{classe.id}_{horodatage}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["Élève", "Absences justifiées", "Absences injustifiées", "Retards"])
    for eleve, j, ij, r in lignes:
        ws.append([eleve.nom_complet, j, ij, r])
    wb.save(chemin_excel)

    rapport = RapportGenere(
        type="absences", titre=titre, fichier_pdf=chemin_pdf, fichier_excel=chemin_excel
    )
    db.session.add(rapport)
    db.session.commit()
    return rapport


def generer_rapport_discipline(cycle):
    lignes = []
    for snapshot in cycle.snapshots:
        infractions = InfractionMineure.query.filter_by(
            cycle_id=cycle.id, eleve_id=snapshot.eleve_id
        ).all()
        lignes.append((snapshot.eleve, snapshot.points_finaux, infractions))

    titre = f"Discipline du {cycle.date_debut} au {cycle.date_fin}"
    horodatage = _horodatage()

    lignes_html = ""
    for eleve, points, infractions in lignes:
        details = "; ".join(f"{i.type_infraction.libelle} (-{i.type_infraction.points_deduits})" for i in infractions)
        lignes_html += f"<tr><td>{eleve.nom_complet}</td><td>{points}/20</td><td>{details or '-'}</td></tr>"

    html = f"""
    <html><body>
    <h2>Rapport de discipline — {titre}</h2>
    <table border="1" cellpadding="4" cellspacing="0" width="100%">
      <tr><th>Élève</th><th>Points finaux</th><th>Infractions de la période</th></tr>
      {lignes_html}
    </table>
    </body></html>
    """
    chemin_pdf = os.path.join(_dossier_rapports(), f"discipline_{cycle.id}_{horodatage}.pdf")
    _ecrire_pdf(html, chemin_pdf)

    chemin_excel = os.path.join(_dossier_rapports(), f"discipline_{cycle.id}_{horodatage}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["Élève", "Points finaux", "Infractions"])
    for eleve, points, infractions in lignes:
        details = "; ".join(f"{i.type_infraction.libelle} (-{i.type_infraction.points_deduits})" for i in infractions)
        ws.append([eleve.nom_complet, points, details])
    wb.save(chemin_excel)

    rapport = RapportGenere(
        type="discipline",
        titre=titre,
        fichier_pdf=chemin_pdf,
        fichier_excel=chemin_excel,
        cycle_id=cycle.id,
    )
    db.session.add(rapport)
    db.session.commit()
    return rapport
